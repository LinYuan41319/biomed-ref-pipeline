from __future__ import annotations

import re
from pathlib import Path
from typing import Mapping


def load_jcr_records(path: str | Path, sheet_name: str | None = None) -> dict[str, dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read JCR workbooks") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        columns = _detect_columns(header)

        records: dict[str, dict[str, str]] = {}
        for row in rows:
            journal = _cell(row, columns["journal"])
            if not journal:
                continue
            record = {
                "jcr_journal": journal,
                "jcr_quartile": _cell(row, columns.get("quartile")),
                "jcr_5yr_if": _cell(row, columns.get("impact")),
                "jcr_2024_if": _cell(row, columns.get("impact_alt")),
            }
            records[normalize_journal(journal)] = record
        return records
    finally:
        workbook.close()


def screen_metadata_rows(
    metadata_rows: list[Mapping[str, str]],
    jcr_records: Mapping[str, Mapping[str, str]],
    *,
    min_if: float = 5.0,
    allowed_quartiles: tuple[str, ...] = ("Q1", "Q2"),
) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for row in metadata_rows:
        match = _match_journal(row, jcr_records)
        impact = _float(match.get("jcr_5yr_if", "") or match.get("jcr_2024_if", "")) if match else None
        quartile = (match.get("jcr_quartile", "") if match else "").upper()
        priority = bool(match) and quartile in allowed_quartiles and impact is not None and impact >= min_if
        output.append(
            {
                **{str(k): str(v) for k, v in row.items()},
                "jcr_match": "yes" if match else "no",
                "jcr_journal": match.get("jcr_journal", "") if match else "",
                "jcr_quartile": quartile,
                "jcr_5yr_if": match.get("jcr_5yr_if", "") if match else "",
                "jcr_2024_if": match.get("jcr_2024_if", "") if match else "",
                "priority_q1q2_if5": "yes" if priority else "no",
            }
        )
    return output


def normalize_journal(value: str) -> str:
    value = value.lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "", value)
    return value


def _match_journal(row: Mapping[str, str], records: Mapping[str, Mapping[str, str]]) -> Mapping[str, str] | None:
    candidates = [row.get("journal", ""), row.get("iso_abbreviation", "")]
    for candidate in candidates:
        key = normalize_journal(str(candidate))
        if key in records:
            return records[key]
    return None


def _detect_columns(header: list[str]) -> dict[str, int]:
    normalized = {normalize_journal(name): idx for idx, name in enumerate(header)}
    columns = {
        "journal": _first(normalized, ["journalname", "journal", "fulljournalname", "journaltitle"]),
        "quartile": _first(normalized, ["jifquartile", "quartile", "jcrquartile"]),
        "impact": _first(normalized, ["5yearjif", "5yearif", "fiveyearjif", "fiveyearimpactfactor"]),
        "impact_alt": _first(normalized, ["jif2024", "2024jif", "journalimpactfactor", "jif"]),
    }
    if columns["journal"] is None:
        raise ValueError(f"Could not detect journal column from header: {header}")
    return {key: value for key, value in columns.items() if value is not None}


def _first(normalized: Mapping[str, int], names: list[str]) -> int | None:
    for name in names:
        if name in normalized:
            return normalized[name]
    return None


def _cell(row: tuple[object, ...], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    return "" if value is None else str(value).strip()


def _float(value: str) -> float | None:
    try:
        return float(str(value).strip())
    except ValueError:
        return None
